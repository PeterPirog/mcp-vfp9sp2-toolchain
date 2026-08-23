from __future__ import annotations

import csv
import json
import re
from collections.abc import Iterator, Mapping
from pathlib import Path
from typing import Any

JSON_READ_CHUNK = 1024 * 1024
OVERFLOW_MARKER = re.compile(r"^\[\[DBFBRIDGE_OVERFLOW:(\d+)\]\]$")


def discover_inputs(source: Path, input_format: str) -> list[Path]:
    suffix = f".{input_format.lower()}"
    candidates = [source] if source.is_file() else list(source.rglob(f"*{suffix}"))
    excluded_names = {
        "migration_report.jsonl",
        "migration_report.csv",
        "verification_report.json",
        "conversion_quality_report.jsonl",
        "conversion_checksums.json",
        "reconstruction_report.jsonl",
    }
    return sorted(
        path
        for path in candidates
        if path.is_file()
        and path.suffix.lower() == suffix
        and path.name not in excluded_names
        and not path.stem.endswith("_schema")
        and not path.stem.endswith(".deleted")
    )


def schema_path_for(data_path: Path) -> Path:
    return data_path.with_name(f"{data_path.stem}_schema.json")


def load_schema(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as infile:
        schema = json.load(infile)
    if schema.get("schema_format") != "dbfbridge-vfp-table-schema":
        raise ValueError(f"Unsupported or missing schema_format in {path.name}.")
    if schema.get("schema_version") != 1:
        raise ValueError(
            f"Unsupported schema_version in {path.name}: {schema.get('schema_version')}"
        )
    fields = schema.get("fields")
    if not isinstance(fields, list) or not fields:
        raise ValueError(f"Schema {path.name} has no field definitions.")
    return schema


def iter_records(
    path: Path, input_format: str, schema: Mapping[str, Any]
) -> Iterator[dict[str, Any]]:
    if input_format == "jsonl":
        yield from _iter_jsonl(path)
    elif input_format == "json":
        yield from _iter_json_array(path)
    elif input_format == "csv":
        yield from _iter_csv(path, schema)
    elif input_format == "xlsx":
        yield from _iter_xlsx(path)
    else:
        raise ValueError(f"Unsupported input format: {input_format}")


def _iter_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as infile:
        for line_number, line in enumerate(infile, start=1):
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path.name}, JSONL line {line_number}: {exc.msg}") from exc
            yield _require_object(record, path, line_number)


def _iter_json_array(path: Path) -> Iterator[dict[str, Any]]:
    decoder = json.JSONDecoder()
    with path.open("r", encoding="utf-8-sig", newline="") as infile:
        buffer = ""
        offset = 0
        eof = False

        def fill() -> bool:
            nonlocal buffer, offset, eof
            if eof:
                return False
            chunk = infile.read(JSON_READ_CHUNK)
            if not chunk:
                eof = True
                return False
            buffer = buffer[offset:] + chunk
            offset = 0
            return True

        fill()
        while True:
            while offset < len(buffer) and buffer[offset].isspace():
                offset += 1
            if offset < len(buffer):
                break
            if not fill():
                raise ValueError(f"{path.name}: empty JSON input.")
        if buffer[offset] != "[":
            raise ValueError(f"{path.name}: top-level JSON value must be an array.")
        offset += 1
        record_number = 0
        expect_value = True
        while True:
            while True:
                while offset < len(buffer) and buffer[offset].isspace():
                    offset += 1
                if offset < len(buffer) or not fill():
                    break
            if offset >= len(buffer):
                raise ValueError(f"{path.name}: unterminated JSON array.")
            if buffer[offset] == "]":
                offset += 1
                break
            if not expect_value:
                if buffer[offset] != ",":
                    raise ValueError(f"{path.name}: expected ',' between JSON records.")
                offset += 1
                expect_value = True
                continue
            while True:
                try:
                    value, end = decoder.raw_decode(buffer, offset)
                    break
                except json.JSONDecodeError as exc:
                    if eof:
                        raise ValueError(
                            f"{path.name}, record {record_number + 1}: {exc.msg}"
                        ) from exc
                    if not fill():
                        continue
            offset = end
            record_number += 1
            yield _require_object(value, path, record_number)
            expect_value = False
        while True:
            while offset < len(buffer) and buffer[offset].isspace():
                offset += 1
            if offset < len(buffer):
                raise ValueError(f"{path.name}: trailing data after JSON array.")
            if not fill():
                return


def _iter_csv(path: Path, schema: Mapping[str, Any]) -> Iterator[dict[str, Any]]:
    fields = {str(field["name"]): field for field in schema["fields"]}
    with path.open("r", encoding="utf-8-sig", newline="") as infile:
        reader = csv.DictReader(infile)
        if reader.fieldnames is None:
            raise ValueError(f"{path.name}: CSV header is missing.")
        for _row_number, row in enumerate(reader, start=2):
            yield {
                name: _parse_csv_cell(value, fields.get(name))
                for name, value in row.items()
                if name is not None
            }


def _parse_csv_cell(value: str | None, field: Mapping[str, Any] | None) -> Any:
    if value is None or value == "":
        return None
    dbf_type = str(field.get("dbf_type")) if field else ""
    if dbf_type in {"C", "V", "D", "T", "@", "M", "G", "P", "Q", "W"}:
        if value.startswith('"'):
            try:
                decoded = json.loads(value)
            except json.JSONDecodeError:
                return value
            if isinstance(decoded, str):
                return decoded
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def _iter_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("XLSX reconstruction requires openpyxl>=3.1.5.") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        overflow = _read_overflow_values(workbook)
        data_sheets = sorted(
            (name for name in workbook.sheetnames if name.startswith("Dane_")),
            key=lambda name: int(name.rsplit("_", 1)[1]),
        )
        if not data_sheets:
            raise ValueError(f"{path.name}: XLSX has no Dane_* worksheets.")
        for sheet_name in data_sheets:
            rows = workbook[sheet_name].iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            names = [str(value) if value is not None else "" for value in header]
            for row in rows:
                record: dict[str, Any] = {}
                for name, value in zip(names, row, strict=False):
                    if not name:
                        continue
                    if isinstance(value, str):
                        marker = OVERFLOW_MARKER.match(value)
                        if marker:
                            overflow_id = int(marker.group(1))
                            if overflow_id not in overflow:
                                raise ValueError(
                                    f"{path.name}: missing overflow value {overflow_id}."
                                )
                            value = overflow[overflow_id]
                    record[name] = value
                yield record
    finally:
        workbook.close()


def _read_overflow_values(workbook: Any) -> dict[int, str]:
    chunks: dict[int, list[tuple[int, int, str]]] = {}
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("Dlugie_teksty_"):
            continue
        rows = workbook[sheet_name].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        names = [str(value) if value is not None else "" for value in header]
        indexes = {name: index for index, name in enumerate(names)}
        required = {"overflow_id", "part", "parts", "text"}
        if not required.issubset(indexes):
            raise ValueError(f"Worksheet {sheet_name} has an invalid overflow header.")
        for row in rows:
            overflow_id = int(row[indexes["overflow_id"]])
            part = int(row[indexes["part"]])
            parts = int(row[indexes["parts"]])
            text = str(row[indexes["text"]] or "")
            chunks.setdefault(overflow_id, []).append((part, parts, text))
    result: dict[int, str] = {}
    for overflow_id, parts in chunks.items():
        parts.sort()
        expected = parts[0][1]
        if len(parts) != expected or [part for part, _, _ in parts] != list(range(1, expected + 1)):
            raise ValueError(f"Overflow value {overflow_id} is incomplete.")
        result[overflow_id] = "".join(text for _, _, text in parts)
    return result


def _require_object(value: Any, path: Path, number: int) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{path.name}, record {number}: expected a JSON object.")
    return value
